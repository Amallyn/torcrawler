# -*- coding: utf-8 -*-

u"""Workbook load/save the Crawler state

Life cycle: see main
"""

__usage__ = u"""Usage: python3 workbook.py [options] [url]

Options:
  -u ..., --url=...       Website URL to crawl
  -h, --help              show this help
  -p, --path              www path where files were stored eg. /var/www
  -d                      show debugging information

Examples:
  python3 workbook.py --url='https://www.nytimes3xbfgragh.onion/' --path='/var/www'
"""

__author__ = u"M0t13y"
__version__ = u"$Revision: 0.01 $"
__date__ = u"$Date: 2021/05/01 17:00:00 $"
__copyright__ = u"Copyright [" + __author__ + "]"
__license__ = u"Licensed under the Apache License, Version 2.0"

import os

# url parser
from urllib.parse import urlsplit

import openpyxl 
from openpyxl.styles import colors
from openpyxl.styles import Font, Color

from datetime import datetime

from settings import WORKBOOK, HTML_SUBDIR

from link import WeightedLink

class CrawlWorkbook():
    """
    Crawl Workbook class
    
    Save/Load progress of a crawl frontier
    Life cycle: see main
    """
    file_name = u''
    html_pages_path = u''
    # workbook
    wb = None

    weighted_links = set()
    weighted_links_done = []
    ignore_seeds = []
    ignored_pages = set()
    
    def __init__(self, path='/var/www', url='http://localhost'):
      """
      Init with file_name
      """
      self.file_name = CrawlWorkbook.wb_filename(path, url)
      self.html_pages_path = CrawlWorkbook.wb_html_pages_path(path, url)

    def __unicode__(self):
        return self.file_name

    @staticmethod  
    def wb_filename(path='/var/www', url='http://localhost'):
        """
        Return the workbook file name from www path and url
        """
        # /var/www => /var/www/html
        html_path = os.path.join(path, HTML_SUBDIR)
        # http://localhost => localhost
        netloc = urlsplit(url).netloc
        # /var/www/html/localhost.xlsx
        file_name = os.path.join(html_path, urlsplit(url).netloc) + WORKBOOK['FILE_EXT']
        return file_name

    @staticmethod  
    def wb_html_pages_path(path='/var/www', url='http://localhost'):
        """
        Return the workbook file name from www path and url
        
        /! Useless in Workbook for now
        """
        # /var/www => /var/www/html
        html_path = os.path.join(path, HTML_SUBDIR)
        # http://localhost => localhost
        netloc = urlsplit(url).netloc
        # /var/www/html/localhost
        html_pages_path = os.path.join(html_path, netloc)
        return html_pages_path

    def ws_load_weighted_links(self, worksheet=None):
        """
        Return weighted links from worksheet
        """
        weighted_links = []
        if worksheet is None:
            return None
        else:
            # Using values_only
            for row in worksheet.iter_rows(min_row=WORKBOOK['crawler']['worksheet']['weightedlink']['STARTING_ROW'], values_only=True):
                weighted_link = WeightedLink(url = row[WORKBOOK['crawler']['worksheet']['weightedlink']['URL_COL']],
                      title = row[WORKBOOK['crawler']['worksheet']['weightedlink']['TITLE_COL']],
                      date = row[WORKBOOK['crawler']['worksheet']['weightedlink']['DATE_COL']],
                      weight = row[WORKBOOK['crawler']['worksheet']['weightedlink']['WEIGHT_COL']],
                      notes = row[WORKBOOK['crawler']['worksheet']['weightedlink']['NOTES_COL']])
                if weighted_link.title is None:
                    weighted_link.title = ''
                if weighted_link.date is None:
                    weighted_link.date = ''
                if weighted_link.weight is None:
                    weighted_link.weight = ''
                if weighted_link.notes is None:
                    weighted_link.notes = ''
                #print(weighted_link)
                weighted_links.append(weighted_link)
            #print(weighted_links)
            return weighted_links

    def ws_load_links(self, worksheet=None):
        """
        Return weighted links from worksheet
        """
        links = []
        if worksheet is None:
            return None
        else:
            # Using values_only
            for row in worksheet.iter_rows(min_row=WEIGHTED_LINK_STARTING_ROW, values_only=True):
                link = row[WEIGHTED_LINK_URL]
                links.append(link)
            #print(weighted_links)
            return links

    def wb_open(self):
        """
        Load or create a workbook if it does not exist
        
        Return [weighted_links_done, weighted_links]
        """
        try:
            # open if already exist
            self.wb = openpyxl.load_workbook(self.file_name)
            
            self.weighted_links_done = self.ws_load_weighted_links(self.wb[WORKBOOK['crawler']['worksheet']['crawledpages']['TITLE']])
            self.weighted_links = self.ws_load_weighted_links(self.wb[WORKBOOK['crawler']['worksheet']['tocrawlpages']['TITLE']])
            self.ignore_seeds = self.ws_load_weighted_links(self.wb[WORKBOOK['crawler']['worksheet']['ignoreseeds']['TITLE']])
            self.ignored_pages = self.ws_load_weighted_links(self.wb[WORKBOOK['crawler']['worksheet']['ignoredpages']['TITLE']])
            print('--- open ---')
            print(self.file_name)
            print('crawled pages:', len(self.weighted_links_done))
            print('to crawl pages:', len(self.weighted_links))
            print('ignore seeds:', len(self.ignore_seeds))
            print('ignored pages:', len(self.ignored_pages))
            print('--- loaded ---')
        except IOError:
            # create
            self.wb_create()
            return None

    def ws_fill_empty(self, ws):
        """
        See ws_init
        """
        # Empty firt row
        ws.append(['', '', '', '', '', ''])
        # freeze first 3 rows and first column
        ws.freeze_panes = "B4"
        
        # categories row
        ws.append(['', 'Url', 'Title', 'Date', 'Weight', 'Notes'])
        
        # set categories style
        categories_row = ws.row_dimensions[2]
        categories_row.font = Font(color="00339966", bold=True)
        
        # columns width
        ws.column_dimensions['B'].width = 50.0
        ws.column_dimensions['C'].width = 50.0
        ws.column_dimensions['D'].width = 20.0
        ws.column_dimensions['F'].width = 50.0
    
        # openpyxl row_max bugs when empty strings
        ws.append([' ', ' ', ' ', ' ', ' ', ' '])
        
    def ws_init(self, worksheet_name, worksheet_num):
        """
        Inits a worksheet with default columns
        
        Workbook should be opened
        Does not save the workbook
        """
        ws = self.wb.create_sheet(worksheet_name, worksheet_num)
        self.ws_fill_empty(ws)
    
    def wb_create(self):
        """
        Create a workbook
        """
        self.wb = openpyxl.Workbook()
        print('--- create ---')
        print(self.file_name)
        for wsname in ['crawledpages', 'tocrawlpages', 'ignoreseeds', 'ignoredpages']:
            print(WORKBOOK['crawler']['worksheet'][wsname]['TITLE'])
            self.ws_init(WORKBOOK['crawler']['worksheet'][wsname]['TITLE'],
                         -1)
                        #WORKBOOK['crawler']['worksheet'][wsname]['INDEX'])
    
        self.wb_save() 
        print('--- save ---')

    def ws_writeln(self, worksheet_name, weighted_link=WeightedLink('', '', '2020-01-31 00:00:00', 0, '')):
        """
        Add a line to the workboot worksheet
        
        Workbook should already be opened
        /!\ Does not save
        """
       
        ws = self.wb[worksheet_name]
        ws.append([ '', weighted_link.url, weighted_link.title,
                  weighted_link.date,
                  weighted_link.weight,
                  weighted_link.notes])

    def ws_appendrows(self, worksheet_name, weighted_links):
        """
        Write links to worksheet
        
        Workbook should already be opened
        /! Parameter: link is absolute  eg.: 'https://www.apple.com/section/useless_page'
        """

        ws = self.wb[worksheet_name]
        for weighted_link in weighted_links:
            ws.append([ '', weighted_link.url, weighted_link.title,
                  weighted_link.date,
                  weighted_link.weight,
                  weighted_link.notes])
        self.wb_save() 

    def ws_writerows(self, worksheet_name, weighted_links):
        """
        Write links to worksheet, Empty sheet first
        
        Workbook should already be opened
        /! Parameter: link is absolute  eg.: 'https://www.apple.com/section/useless_page'
        """

        ws = self.wb[worksheet_name]

        row_min = WORKBOOK['crawler']['worksheet']['weightedlink']['STARTING_ROW']
        row_max = ws.max_row
        #print(worksheet_name, row_max)
        #print('row min:', row_min, 'row max', row_max)
        # Empty sheet first if not empty
        if row_max > row_min:
            ws.delete_rows(row_min, row_max)
        for weighted_link in weighted_links:
            #print(weighted_link.url)
            ws.append([ '', weighted_link.url, weighted_link.title,
                  weighted_link.date,
                  weighted_link.weight,
                  weighted_link.notes])
        self.wb_save() 

    def wb_save(self):
        """
        Save the workbook
        """    
        self.wb.save(self.file_name)

class Usage(Exception):
    def __init__(self, msg):
        self.msg = msg
        
def main(argv=None):
    """
    main
    """
    import getopt
    if argv is None:
        argv = sys.argv

    try:
        try:                                
            opts, args = getopt.getopt(argv, "hup:d", ["help", "url=", "path="])
        except getopt.error as msg:
             raise Usage(msg)            
        url = u'http://localhost'
        path = u'/var/www'
        for opt, arg in opts:
            if opt in ("-h", "--help"):
                print(__usage__)                   
                sys.exit()                  
            elif opt == '-d':
                global _debug               
                _debug = 1                  
            elif opt in ("-u", "--url"):
                url = arg               
            elif opt in ("-p", "--path"):
                path = arg               
    except Usage as err:
        print >>sys.stderr, err.msg
        print >>sys.stderr, "for help use --help"
        return 2

    if len(argv)<2:
        print("Arguments error")
        print(__usage__)
        sys.exit(2)

    w_l_00 = WeightedLink(u'https://www.apple.xlsx/',
                      u'New York Times',
                      datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                      1024,
                      u'bla bla')
                      
    w_l_01 = WeightedLink(u'https://www.apple.xlsx/mehpage',
                      u'New York Times - Meh',
                      datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                      512,
                      u'meh')
                      
    w_l_02 = WeightedLink(u'https://www.apple.xlsx/megapage',
                      u'New York Times - mega',
                      datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                      2048,
                      u'mega bla bla')
    w_l_03 = WeightedLink(u'https://www.apple.xlsx/useless',
                      u'New York Times - useless',
                      datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                      2048,
                      u'mega no no')
    w_l_04 = WeightedLink(u'https://www.apple.xlsx/nopage',
                      u'New York Times - mega',
                      datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                      2048,
                      u'no no')
    
    cwb = CrawlWorkbook(path, url)
    #print(cwb.file_name)
    #print(cwb.html_pages_path)
    cwb.wb_open()
    print(cwb.weighted_links_done)
    print(cwb.weighted_links)
    print(cwb.ignore_seeds)
    print(cwb.ignored_pages)

    wbwsname = WORKBOOK['crawler']['worksheet']['crawledpages']['TITLE']
    cwb.ws_writeln(wbwsname, w_l_00)
    cwb.wb_save()

    wbwsname = WORKBOOK['crawler']['worksheet']['tocrawlpages']['TITLE']
    cwb.ws_appendrows(wbwsname, [ w_l_01, w_l_02 ])

    wbwsname = WORKBOOK['crawler']['worksheet']['ignoredpages']['TITLE']
    cwb.ws_writerows(wbwsname, [ w_l_03, w_l_04 ])
                    
if __name__ == "__main__":
    import sys
    print("\n" + __doc__ + "\n" + __copyright__ + "\n" + __license__ +"\n" )
    main(sys.argv[1:])

