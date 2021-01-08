# -*- coding: utf-8 -*-

u"""crawl_serialize

crawl serialize
"""

__author__ = u"M0t13y"
__version__ = u"$Revision: 0.01 $"
__date__ = u"$Date: 2021/05/01 17:00:00 $"
__copyright__ = u"Copyright [" + __author__ + "]"
__license__ = u"Licensed under the Apache License, Version 2.0"

import openpyxl 
from openpyxl.styles import colors
from openpyxl.styles import Font, Color

from datetime import datetime

from weighted_link import WeightedLink

WORKSHEET_CRAWLED_PAGES = u'Crawled pages'
WORKSHEET_PAGES_TO_CRAWL = u'Pages to Crawl'

# WeightedLink fields columns in a Worksheet 
WEIGHTED_LINK_URL = 1
WEIGHTED_LINK_TITLE = 2
WEIGHTED_LINK_DATE = 3
WEIGHTED_LINK_WEIGHT = 4
WEIGHTED_LINK_NOTES = 5

WEIGHTED_LINK_STARTING_ROW = 4

class CrawlWorkbook():
    """
    Crawl Workbook class
    
    Save/Load progress of a crawl frontier
    Life cycle (see main):
      cw = CrawlWorkbook(u'www.apple.xlsx')
      cw.wb_open()
      cw.wb_writeln(WORKSHEET_CRAWLED_PAGES, weighted_link_01)
      cw.wb_save()
      cw.wb_reset(WORKSHEET_PAGES_TO_CRAWL)
      for weightedlink in weightedlinks:
          cw.wb_writeln(WORKSHEET_PAGES_TO_CRAWL, weightedlink)
      cw.wb_save()
    """
    file_name = u''
    # workbook
    wb = None
    
    def __init__(self, file_name=u''):
      """
      Init with file_name
      """
      self.file_name = file_name

    def __unicode__(self):
        return self.file_name
      
    def ws_load_weighted_links(self, worksheet=None):
        """
        Return weighted links from worksheet
        """
        weighted_links = []
        if worksheet is None:
            return None
        else:
            # Using values_only
            for row in worksheet.iter_rows(min_row=WEIGHTED_LINK_STARTING_ROW, values_only=True):
                weighted_link = WeightedLink(url = row[WEIGHTED_LINK_URL],
                      title = row[WEIGHTED_LINK_TITLE],
                      date = row[WEIGHTED_LINK_DATE],
                      weight = row[WEIGHTED_LINK_WEIGHT],
                      notes = row[WEIGHTED_LINK_NOTES])
                if weighted_link.notes is None:
                    weighted_link.notes = ''
                #print(weighted_link)
                weighted_links.append(weighted_link)
            #print(weighted_links)
            return weighted_links

    def wb_open(self):
        """
        Load or create a workbook if it does not exist
        
        Return [weighted_links_done, weighted_links]
        """
        weighted_links = []
        weighted_links_done = []
        try:
            # open if already exist
            self.wb = openpyxl.load_workbook(self.file_name)
            
            weighted_links_done = self.ws_load_weighted_links(self.wb[WORKSHEET_CRAWLED_PAGES])
            weighted_links = self.ws_load_weighted_links(self.wb[WORKSHEET_PAGES_TO_CRAWL])
            return [weighted_links_done, weighted_links]
        except IOError:
            # create
            self.wb_create()
            return None

    def ws_init(self, worksheet_name, worksheet_num):
        """
        Inits a worksheet with default columns
        
        Workbook should be opened
        Does not save the workbook
        """
        ws = self.wb.create_sheet(worksheet_name, worksheet_num)
        # Empty firt row
        ws.append(['', '', '', '', '', ''])
        # freeze first 3 rows and first column
        ws.freeze_panes = "B4"
        
        # categories row
        ws.append(['', 'Url', 'Title', 'Date', 'Weight', 'Notes'])
        
        # set categories style
        categories_row = ws.row_dimensions[2]
        categories_row.font = Font(color="00339966", bold=True)
        
        # columns width
        ws.column_dimensions['B'].width = 50.0
        ws.column_dimensions['C'].width = 50.0
        ws.column_dimensions['D'].width = 20.0
        ws.column_dimensions['F'].width = 50.0
    
        ws.append(['', '', '', '', '', ''])
    
    def wb_create(self):
        """
        Create a workbook
        """
        self.wb = openpyxl.Workbook()
        # Workbook's active sheet: Crawled pages
        self.ws_init(WORKSHEET_CRAWLED_PAGES, 0)
        self.ws_init(WORKSHEET_PAGES_TO_CRAWL, 1)
    
        self.wb.save(self.file_name) 

    def ws_writeln(self, worksheet_name, weighted_link=WeightedLink('', '', '2020-01-31 00:00:00', 0, '')):
        """
        Add a line to the workboot worksheet
        
        Workbook should already be opened
        /!\ Does not save
        """
       
        ws = self.wb[worksheet_name]
        ws.append([ '', weighted_link.url, weighted_link.title,
                  weighted_link.date,
                  weighted_link.weight,
                  weighted_link.notes])

    def wb_save(self):
        """
        Save the workbook
        """    
        self.wb.save(self.file_name)

    def ws_reset(self, worksheet_name):
        """
        Write a complete workbook worksheet
        
        Workbook should already be opened
        """
    
        # select & delete worksheet
        #index = self.wb.worksheets.index(worksheet_name)
        index = 1
        ws = self.wb[worksheet_name]
        self.wb.remove(ws)
    
        # init
        self.ws_init( worksheet_name, index)
            
        #saving the file
        self.wb_save()
            
if __name__ == "__main__":
    print("\n" + __doc__ + "\n" + __copyright__ + "\n" + __license__ +"\n" )
    
    w_l_00 = WeightedLink(u'https://www.nytimes3xbfgragh.onion/',
                      u'New York Times',
                      datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                      1024,
                      u'bla bla')
                      
    w_l_01 = WeightedLink(u'https://www.nytimes3xbfgragh.onion/mehpage',
                      u'New York Times - Meh',
                      datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                      512,
                      u'meh')
                      
    w_l_02 = WeightedLink(u'https://www.nytimes3xbfgragh.onion/megapage',
                      u'New York Times - mega',
                      datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                      2048,
                      u'mega bla bla')
    
    crawl_wb = CrawlWorkbook('www.nytimes3xbfgragh.onion.xlsx')
    crawl_wb.wb_open()
    #crawl_wb.ws_writeln(WORKSHEET_CRAWLED_PAGES, w_l_00)
    #crawl_wb.wb_save()
    #crawl_wb.ws_reset(WORKSHEET_PAGES_TO_CRAWL)
    #for w_l in [ w_l_01, w_l_02 ]:
    #    crawl_wb.ws_writeln(WORKSHEET_PAGES_TO_CRAWL, w_l)
    #crawl_wb.wb_save()
    
    